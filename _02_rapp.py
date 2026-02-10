import pandas as pd
import re
import _04_pdf_utils as pdf_utils
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

def executer_rapprochement(data_banque, data_compta, data_etat_prec=None, date_rapprochement=None):
    """
    Exécute le rapprochement bancaire entièrement en mémoire.
    
    Args:
        data_banque: DataFrame ou file-like object (Excel)
        data_compta: DataFrame ou file-like object (Excel)
        data_etat_prec: DataFrame ou file-like object (Excel) (Optionnel)
        date_rapprochement: Date/Datetime/String
        
    Returns:
        tuple: (excel_bytes: io.BytesIO, pdf_bytes: bytes, stats: dict)
    """
    
    # Helpers pour la gestion des dates
    def format_date_val(val):
        if pd.isna(val): return ""
        try:
            return pd.to_datetime(val, dayfirst=True).strftime('%d/%m/%Y')
        except:
            return val
    
    def get_date_obj(val):
        if pd.isna(val): return pd.Timestamp.min
        try:
            return pd.to_datetime(val, dayfirst=True)
        except:
            return pd.Timestamp.min

    def load_data(data, header=0):
        if data is None: return None
        if isinstance(data, pd.DataFrame):
            return data.copy()
        return pd.read_excel(data, header=header)

    # Chargement
    try:
        df_banque_raw = load_data(data_banque) # Gardé pour recherche solde
        df_compta_raw = load_data(data_compta, header=0) # Gardé pour recherche solde
        
        df_banque = df_banque_raw.copy()
        df_compta = df_compta_raw.copy()
    except Exception as e:
        raise ValueError(f"Erreur lors du chargement des données : {e}")

    # Normalisation des noms de colonnes (Débit -> debit, Crédit -> credit)
    df_banque.rename(columns=lambda x: str(x).lower().replace('é', 'e'), inplace=True)
    
    # Nettoyage : Suppression de la ligne "Solde précédent" si présente
    for col in df_banque.select_dtypes(include=['object']).columns:
        df_banque = df_banque[~df_banque[col].astype(str).str.contains("Solde précédent", case=False, na=False)]

    # Normalisation Compta
    df_compta.rename(columns=lambda x: str(x).lower().replace('é', 'e'), inplace=True)

    # Nettoyage des données (remplacement des NaN par 0 pour calculs)
    cols_montants = ['debit', 'credit']
    # On s'assure que les colonnes existent
    for col in cols_montants:
        if col not in df_banque.columns: df_banque[col] = 0
        if col not in df_compta.columns: df_compta[col] = 0
            
    df_banque[cols_montants] = df_banque[cols_montants].fillna(0)
    df_compta[cols_montants] = df_compta[cols_montants].fillna(0)

    # --- NETTOYAGE JOURNAL : SUPPRESSION DES ANNULATIONS (e.g. 1500 et -1500) ---
    def get_indices_annulation(df, col):
        indices = []
        pos_map = {}
        # On mappe les positifs
        for idx, val in df[col][df[col] > 0].items():
            v = round(val, 4)
            pos_map.setdefault(v, []).append(idx)
        # On cherche les correspondances avec les négatifs
        for idx, val in df[col][df[col] < 0].items():
            target = round(abs(val), 4)
            if target in pos_map and pos_map[target]:
                indices.append(idx)
                indices.append(pos_map[target].pop(0))
        return indices

    # ----------------------------------------------------------------------------------
    # FONCTION OPERATION ANNULEE (NOUVEAU)
    # Vérifie le restant des transactions non pointées du relevé en opposant débit et crédit.
    # Si montant identique et libellé similaire (regex), on supprime.
    # ----------------------------------------------------------------------------------
    def operation_annulée(df):
        if df.empty: return df
        
        # On travaille sur une copie pour les calculs, mais on veut renvoyer le meme type de DF
        # On va identifier les indices à supprimer
        to_drop = []
        
        # Séparation (vues)
        # S'assurer que les colonnes existent
        if 'debit' not in df.columns or 'credit' not in df.columns:
            return df
            
        debits = df[df['debit'] > 0]
        credits = df[df['credit'] > 0]
        
        # Helper de similarité basé sur les NUMÉROS (ex: N° de chèque)
        def check_similarity(lib1, lib2):
            if not isinstance(lib1, str) or not isinstance(lib2, str):
                return False
            
            # Extraction des séquences de chiffres
            # On cherche des identifiants numériques.
            # On utilise une regex pour trouver tous les nombres (suite de digits)
            nums1 = set(re.findall(r'\d+', lib1))
            nums2 = set(re.findall(r'\d+', lib2))
            
            # Filtrage des nombres non significatifs
            # On ignore les nombres trop courts (ex: 1 ou 2 chiffres) qui pourraient être des jours ou mois isolés
            # On garde les nombres de longueur >= 3 (ex: 100, 2023, 495839...)
            # Cela permet de matcher des numéros de chèque (souvent 7 chiffres) ou des années/références
            nums1 = {n for n in nums1 if len(n) >= 3}
            nums2 = {n for n in nums2 if len(n) >= 3}
            
            # Intersection : Y a-t-il au moins un numéro commun ?
            common = nums1.intersection(nums2)
            
            return len(common) > 0

        # Set des indices utilisés côté crédit pour éviter d'utiliser le meme crédit pour 2 débits
        used_credit_indices = set()
        
        for idx_d, row_d in debits.iterrows():
            amount = row_d['debit']
            lib_d = str(row_d.get('libelle', ''))
            
            # Candidats crédits (Même montant exact)
            candidates = credits[
                (credits['credit'] == amount) & 
                (~credits.index.isin(used_credit_indices))
            ]
            
            if candidates.empty:
                continue
                
            best_match_idx = None
            
            for idx_c, row_c in candidates.iterrows():
                lib_c = str(row_c.get('libelle', ''))
                
                # Check regex similarity
                if check_similarity(lib_d, lib_c):
                    best_match_idx = idx_c
                    break
            
            if best_match_idx is not None:
                # On marque les deux pour suppression
                to_drop.append(idx_d)
                to_drop.append(best_match_idx)
                used_credit_indices.add(best_match_idx)
                
        if to_drop:
            print(f"  --> Opérations annulées détectées et supprimées : {len(to_drop)//2} paires.")
            return df.drop(to_drop), df.loc[to_drop]
            
        return df, pd.DataFrame(columns=df.columns)

    drop_d = get_indices_annulation(df_compta, 'debit')
    drop_c = get_indices_annulation(df_compta, 'credit')
    indices_a_supprimer = list(set(drop_d + drop_c))
    
    if indices_a_supprimer:
        # print(f"Suppression de {len(indices_a_supprimer)} lignes d'annulations dans le journal.")
        df_compta = df_compta.drop(indices_a_supprimer)

    # Listes pour stocker les lignes rapprochées
    indices_banque_ok = []
    indices_compta_ok = []
    
    suspens_etat_prec = []

    # --- LOGIQUE POINTAGE PREALABLE (ETAT PRECEDENT) ---
    if data_etat_prec is not None:
        print(f"Traitement de l'état précédent...")
        try:
            df_etat = load_data(data_etat_prec, header=None)
            
            start_idx = 3 # Ligne 4
            if len(df_etat) > start_idx:
                for idx, row in df_etat.iloc[start_idx:-3].iterrows():
                    val_lib = str(row[1]) if pd.notna(row[1]) else ""
                    if any(x in val_lib.lower() for x in ["total", "totaux", "solde"]):
                        continue
                    
                    if len(row) < 6: continue
                    
                    val_c = row[2] if isinstance(row[2], (int, float)) else 0
                    val_d = row[3] if isinstance(row[3], (int, float)) else 0
                    val_e = row[4] if isinstance(row[4], (int, float)) else 0
                    val_f = row[5] if isinstance(row[5], (int, float)) else 0
                    
                    keep_c, keep_d, keep_e, keep_f = 0, 0, 0, 0
                    
                    # Verification Compta
                    if val_c > 0:
                        match = df_compta[(df_compta['debit'] == val_c) & (~df_compta.index.isin(indices_compta_ok))]
                        if not match.empty: indices_compta_ok.append(match.index[0])
                        else: keep_c = val_c

                    if val_d > 0:
                        match = df_compta[(df_compta['credit'] == val_d) & (~df_compta.index.isin(indices_compta_ok))]
                        if not match.empty: indices_compta_ok.append(match.index[0])
                        else: keep_d = val_d

                    # Verification Banque
                    if val_e > 0:
                        match = df_banque[(df_banque['debit'] == val_e) & (~df_banque.index.isin(indices_banque_ok))]
                        if not match.empty: indices_banque_ok.append(match.index[0])
                        else: keep_e = val_e
                            
                    if val_f > 0:
                        match = df_banque[(df_banque['credit'] == val_f) & (~df_banque.index.isin(indices_banque_ok))]
                        if not match.empty: indices_banque_ok.append(match.index[0])
                        else: keep_f = val_f
                    
                    if any([keep_c, keep_d, keep_e, keep_f]):
                        suspens_etat_prec.append({
                            'raw_date': get_date_obj(row[0]),
                            'date_str': format_date_val(row[0]),
                            'libelle': row[1],
                            'col_C': keep_c, 'col_D': keep_d, 'col_E': keep_e, 'col_F': keep_f
                        })

        except Exception as e:
            print(f"Erreur état précédent : {e}")

    # --- LOGIQUE DE POINTAGE ---
    for idx_b, row_b in df_banque.iterrows():
        # IMPORTANT: Si cette ligne banque a déjà été utilisée (par ex. par l'état précédent), on passe.
        if idx_b in indices_banque_ok:
            continue

        if row_b['debit'] > 0:
            match = df_compta[(df_compta['credit'] == row_b['debit']) & (~df_compta.index.isin(indices_compta_ok))]
            if not match.empty:
                indices_banque_ok.append(idx_b)
                indices_compta_ok.append(match.index[0])
        elif row_b['credit'] > 0:
            match = df_compta[(df_compta['debit'] == row_b['credit']) & (~df_compta.index.isin(indices_compta_ok))]
            if not match.empty:
                indices_banque_ok.append(idx_b)
                indices_compta_ok.append(match.index[0])

    # --- EXTRACTION DES SUSPENS ---
    suspens_banque = df_banque.drop(indices_banque_ok)
    
    # APPLICATION DU FILTRE OPERATION ANNULEE SUR LE RELEVE (SUSPENS)
    # APPLICATION DU FILTRE OPERATION ANNULEE SUR LE RELEVE (SUSPENS)
    suspens_banque, ops_annulees_banque = operation_annulée(suspens_banque)
    
    suspens_compta = df_compta.drop(indices_compta_ok)

    # Préparation Export
    cols_to_drop_banque = [c for c in suspens_banque.columns if 'solde' in str(c).lower()]
    suspens_banque_export = suspens_banque.drop(columns=cols_to_drop_banque)
    
    cols_to_drop_compta = [c for c in suspens_compta.columns if 'solde' in str(c).lower() or 'unnamed' in str(c).lower()]
    suspens_compta_export = suspens_compta.drop(columns=cols_to_drop_compta)

    col_date_compta = next((c for c in suspens_compta_export.columns if 'date' in str(c).lower()), None)
    if col_date_compta:
        try:
            suspens_compta_export[col_date_compta] = pd.to_datetime(suspens_compta_export[col_date_compta], errors='coerce').dt.strftime('%d/%m/%Y')
        except: pass

    cols_to_drop_annulees = [c for c in ops_annulees_banque.columns if 'solde' in str(c).lower()]
    ops_annulees_banque_export = ops_annulees_banque.drop(columns=cols_to_drop_annulees)

    # Création du buffer Excel
    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        suspens_banque_export.to_excel(writer, sheet_name='RELEVE_NON_POINTEE', index=False)
        suspens_compta_export.to_excel(writer, sheet_name='JOURNAL_NON_POINTEE', index=False)
        
        if not ops_annulees_banque_export.empty:
            ops_annulees_banque_export.to_excel(writer, sheet_name='OPERATIONS_ANNULEES', index=False)

    # Ajout feuille Rapprochement avec OpenPyXL
    # Ajout feuille Rapprochement avec OpenPyXL
    excel_buffer.seek(0)
    wb = openpyxl.load_workbook(excel_buffer)
    
    # -----------------------------------------------------------
    # MISE EN FORME DES FEUILLES PRECEDENTES (Nombre #,##0)
    # -----------------------------------------------------------
    accounting_format = '#,##0'
    sheets_to_format = ['RELEVE_NON_POINTEE', 'JOURNAL_NON_POINTEE', 'OPERATIONS_ANNULEES']
    
    for sh_name in sheets_to_format:
        if sh_name in wb.sheetnames:
            ws_tmp = wb[sh_name]
            # On parcourt les colonnes pour trouver Debit/Credit
            # On suppose que les en-têtes sont en ligne 1
            header_cells = list(ws_tmp[1]) # Tuple of cells
            col_indices_to_format = []
            
            for cell in header_cells:
                val = str(cell.value).lower()
                if 'debit' in val or 'credit' in val or 'montant' in val:
                    col_indices_to_format.append(cell.column) # 1-based index
            
            if col_indices_to_format:
                for row in ws_tmp.iter_rows(min_row=2):
                    for cell in row:
                        if cell.column in col_indices_to_format:
                            cell.number_format = accounting_format
    
    wb.save(excel_buffer) # Sauvegarde interne avant de continuer
    
    
    # CALCUL DES SOLDES
    def get_last_solde(df):
        col_solde = next((c for c in df.columns if str(c).lower().strip() == 'solde'), None)
        if col_solde:
            series_valid = df[col_solde].dropna()
            if not series_valid.empty: return series_valid.iloc[-1]
        return 0

    solde_banque = get_last_solde(df_banque_raw)
    solde_compta = get_last_solde(df_compta_raw)

    sheet_name_rapp = "RAPPROCHEMENT"
    if sheet_name_rapp in wb.sheetnames: del wb[sheet_name_rapp]
    ws = wb.create_sheet(sheet_name_rapp)

    # En-têtes, Styles (identique code original)
    ws['A1'] = "Date"; ws['B1'] = "Libellés"; ws['C1'] = "Compte courant"; ws['E1'] = "Relevé bancaire"
    ws['C2'] = "Débit"; ws['D2'] = "Crédit"; ws['E2'] = "Débit"; ws['F2'] = "Crédit"
    ws['B3'] = "Solde à rectifier"; ws['C3'] = solde_compta; ws['F3'] = solde_banque

    ws.merge_cells('A1:A2'); ws.merge_cells('B1:B2'); ws.merge_cells('C1:D1'); ws.merge_cells('E1:F1')
    
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=2, max_col=6):
        for cell in row:
            cell.font = bold_font; cell.alignment = center_align; cell.border = thin_border

    # Ligne 3 styles
    for col in range(1, 7):
        cell = ws.cell(row=3, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left') if col == 2 else center_align

    ws.column_dimensions['B'].width = 40
    for c in ['A','C','D','E','F']: ws.column_dimensions[c].width = 15

    # Collecte Opérations
    col_date_c = next((c for c in suspens_compta.columns if 'date' in str(c).lower()), None)
    col_lib_c = next((c for c in suspens_compta.columns if 'libell' in str(c).lower()), None)
    col_date_b = next((c for c in suspens_banque.columns if 'date' in str(c).lower()), None)
    col_lib_b = next((c for c in suspens_banque.columns if 'libell' in str(c).lower()), None)
    
    all_ops = []
    
    for item in suspens_etat_prec:
        if any([item['col_C'], item['col_D'], item['col_E'], item['col_F']]):
            all_ops.append(item)

    # Compta
    for _, row in suspens_compta.iterrows():
        d_val = row.get('debit', 0); c_val = row.get('credit', 0)
        raw_date = row[col_date_c] if col_date_c else None
        lib = row[col_lib_c] if col_lib_c else ""
        
        if d_val > 0:
            all_ops.append({'raw_date': get_date_obj(raw_date), 'date_str': format_date_val(raw_date), 'libelle': lib, 'col_C': 0, 'col_D': 0, 'col_E': 0, 'col_F': d_val})
        elif c_val > 0:
            all_ops.append({'raw_date': get_date_obj(raw_date), 'date_str': format_date_val(raw_date), 'libelle': lib, 'col_C': 0, 'col_D': 0, 'col_E': c_val, 'col_F': 0})

    # Banque
    for _, row in suspens_banque.iterrows():
        d_val = row.get('debit', 0); c_val = row.get('credit', 0)
        raw_date = row[col_date_b] if col_date_b else None
        lib = row[col_lib_b] if col_lib_b else ""
        
        if d_val > 0:
            all_ops.append({'raw_date': get_date_obj(raw_date), 'date_str': format_date_val(raw_date), 'libelle': lib, 'col_C': 0, 'col_D': d_val, 'col_E': 0, 'col_F': 0})
        elif c_val > 0:
            all_ops.append({'raw_date': get_date_obj(raw_date), 'date_str': format_date_val(raw_date), 'libelle': lib, 'col_C': c_val, 'col_D': 0, 'col_E': 0, 'col_F': 0})

    all_ops.sort(key=lambda x: x['raw_date'])

    current_row = 4
    if not all_ops: current_row += 1 

    for op in all_ops:
        ws[f'A{current_row}'] = op['date_str']
        ws[f'B{current_row}'] = op['libelle']
        if op['col_C'] > 0: ws[f'C{current_row}'] = op['col_C']
        if op['col_D'] > 0: ws[f'D{current_row}'] = op['col_D']
        if op['col_E'] > 0: ws[f'E{current_row}'] = op['col_E']
        if op['col_F'] > 0: ws[f'F{current_row}'] = op['col_F']
        
        for col in range(1, 7):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left') if col == 2 else center_align
        current_row += 1

    # Totaux
    row_total = current_row
    ws[f'B{current_row}'] = "Totaux"
    ws[f'B{current_row}'].font = bold_font; ws[f'B{current_row}'].alignment = Alignment(horizontal='center'); ws[f'B{current_row}'].border = thin_border
    
    # Calcul des sommes de colonnes (Valeurs numériques pour affichage correct dans l'aperçu)
    # Attention: ligne 3 contient les soldes initiaux
    
    # Conversion safe
    def to_float(x):
        try: return float(x)
        except: return 0.0
        
    v_solde_compta = to_float(solde_compta)
    v_solde_banque = to_float(solde_banque)
    
    # Somme des opérations (déjà dans all_ops)
    s_c = sum(to_float(op.get('col_C', 0)) for op in all_ops)
    s_d = sum(to_float(op.get('col_D', 0)) for op in all_ops)
    s_e = sum(to_float(op.get('col_E', 0)) for op in all_ops)
    s_f = sum(to_float(op.get('col_F', 0)) for op in all_ops)
    
    # Totaux ligne (Solde Init + Mouvements)
    # C3 = Solde Compta
    # F3 = Solde Banque
    # D3, E3 vides
    
    t_c = v_solde_compta + s_c
    t_d = s_d
    t_e = s_e
    t_f = v_solde_banque + s_f
    
    ws[f'C{current_row}'] = t_c
    ws[f'D{current_row}'] = t_d
    ws[f'E{current_row}'] = t_e
    ws[f'F{current_row}'] = t_f
    
    for col in ['C', 'D', 'E', 'F']:
        cell = ws[f'{col}{current_row}']
        cell.font = bold_font; cell.border = thin_border; cell.alignment = center_align
    ws[f'A{current_row}'].border = thin_border
    
    current_row += 1
    
    # Solde rectifié
    label_rectif = "Solde rectifié"
    if date_rapprochement:
        try: dstr = date_rapprochement.strftime('%d/%m/%Y')
        except: dstr = str(date_rapprochement)
        label_rectif = f"Solde rectifié au {dstr}"
        
    ws[f'B{current_row}'] = label_rectif
    ws[f'B{current_row}'].font = bold_font; ws[f'B{current_row}'].alignment = Alignment(horizontal='center'); ws[f'B{current_row}'].border = thin_border
    
    # Calcul logiques rectifiés (D - C, etc)
    # IF(D-C>0, D-C, "")
    v_rect_c = (t_d - t_c) if (t_d - t_c) > 0.001 else 0
    v_rect_d = (t_c - t_d) if (t_c - t_d) > 0.001 else 0
    v_rect_e = (t_f - t_e) if (t_f - t_e) > 0.001 else 0
    v_rect_f = (t_e - t_f) if (t_e - t_f) > 0.001 else 0
    
    ws[f'C{current_row}'] = v_rect_c if v_rect_c != 0 else ""
    ws[f'D{current_row}'] = v_rect_d if v_rect_d != 0 else ""
    ws[f'E{current_row}'] = v_rect_e if v_rect_e != 0 else ""
    ws[f'F{current_row}'] = v_rect_f if v_rect_f != 0 else ""
    
    for col in ['A','C','D','E','F']:
        ws[f'{col}{current_row}'].border = thin_border; ws[f'{col}{current_row}'].alignment = center_align; ws[f'{col}{current_row}'].font = bold_font
    
    row_rectifie = current_row
    current_row += 1
    
    # Totaux Généraux
    ws[f'B{current_row}'] = "TOTAUX GENERAUX"
    ws[f'B{current_row}'].font = bold_font; ws[f'B{current_row}'].alignment = Alignment(horizontal='center'); ws[f'B{current_row}'].border = thin_border
    
    ws[f'C{current_row}'] = t_c + v_rect_c
    ws[f'D{current_row}'] = t_d + v_rect_d
    ws[f'E{current_row}'] = t_e + v_rect_e
    ws[f'F{current_row}'] = t_f + v_rect_f
    
    for col in ['C', 'D', 'E', 'F']:
        cell = ws[f'{col}{current_row}']
        cell.font = bold_font; cell.border = thin_border; cell.alignment = center_align
    ws[f'A{current_row}'].border = thin_border

    # Format
    accounting_format = '#,##0'
    for row in range(3, current_row + 1):
        for col in ['C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].number_format = accounting_format

    # Sauvegarde finale en mémoire
    final_excel = io.BytesIO()
    wb.save(final_excel)
    final_excel.seek(0)
    
    # --- PDF GENERATION ---
    t_c = sum(op.get('col_C', 0) for op in all_ops)
    t_d = sum(op.get('col_D', 0) for op in all_ops)
    t_e = sum(op.get('col_E', 0) for op in all_ops)
    t_f = sum(op.get('col_F', 0) for op in all_ops)
    
    try: val_solde_compta = float(solde_compta) if solde_compta else 0
    except: val_solde_compta = 0
    try: val_solde_banque = float(solde_banque) if solde_banque else 0
    except: val_solde_banque = 0
    
    total_C = val_solde_compta + t_c
    total_D = t_d
    total_E = t_e
    total_F = val_solde_banque + t_f
    totals = {'C': total_C, 'D': total_D, 'E': total_E, 'F': total_F}
    
    rect_C, rect_D, rect_E, rect_F = 0, 0, 0, 0
    diff_c = total_C - total_D
    if diff_c < 0: rect_C = abs(diff_c)
    else: rect_D = diff_c
    
    diff_b = total_E - total_F
    if diff_b < 0: rect_E = abs(diff_b)
    else: rect_F = diff_b
    
    solde_rectif = {'label': label_rectif, 'C': rect_C, 'D': rect_D, 'E': rect_E, 'F': rect_F}
    grand_totals = {'C': total_C+rect_C, 'D': total_D+rect_D, 'E': total_E+rect_E, 'F': total_F+rect_F}
    
    ops_for_pdf = [{'date_str':'', 'libelle':'Solde à rectifier', 'col_C':val_solde_compta, 'col_D':0, 'col_E':0, 'col_F':val_solde_banque}] + all_ops
    
    _, pdf_bytes = pdf_utils.generate_pdf_report(ops_for_pdf, totals, solde_rectif, grand_totals, None, date_arrete=str(date_rapprochement) if date_rapprochement else "")

    stats = {
        'suspens_banque': len(suspens_banque),
        'suspens_compta': len(suspens_compta)
    }
    
    return final_excel, pdf_bytes, stats