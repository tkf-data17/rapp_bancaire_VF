import pandas as pd
import os

def executer_rapprochement(path_banque, path_compta, path_etat_prec=None, output_path='documents/Etat_Rapprochement_Automatique.xlsx', date_rapprochement=None):
    
    # Helpers pour la gestion des dates (déplacés en début de fonction pour usage global)
    def format_date_val(val):
        if pd.isna(val): return ""
        try:
            return pd.to_datetime(val, dayfirst=True).strftime('%d/%m/%Y')
        except:
            return val
    
    def get_date_obj(val):
        if pd.isna(val): return pd.Timestamp.min
        try:
            return pd.to_datetime(val, dayfirst=True)
        except:
            return pd.Timestamp.min
    # Chargement avec pandas
    # On force les colonnes Debit et Credit en numérique pour éviter les erreurs de calcul
    df_banque = pd.read_excel(path_banque)
    
    # Nettoyage : Suppression de la ligne "Solde précédent" si présente
    # On cherche dans les colonnes de type texte la mention "Solde précédent" et on exclut ces lignes
    for col in df_banque.select_dtypes(include=['object']).columns:
        df_banque = df_banque[~df_banque[col].astype(str).str.contains("Solde précédent", case=False, na=False)]

    df_compta = pd.read_excel(path_compta, header=1)

    # Normalisation des noms de colonnes pour la compta (Débit -> debit, Crédit -> credit)
    df_compta.rename(columns=lambda x: str(x).lower().replace('é', 'e'), inplace=True)


    # Nettoyage des données (remplacement des NaN par 0)
    cols_montants = ['debit', 'credit']
    df_banque[cols_montants] = df_banque[cols_montants].fillna(0)
    df_compta[cols_montants] = df_compta[cols_montants].fillna(0)

    # Listes pour stocker les lignes rapprochées
    indices_banque_ok = []
    indices_compta_ok = []
    
    suspens_etat_prec = []

    # --- LOGIQUE POINTAGE PREALABLE (ETAT PRECEDENT) ---
    # Structure Etat Prec: A=Date(0), B=Lib(1), C=Debit Compta(2), D=Credit Compta(3), E=Debit Banque(4), F=Credit Banque(5)
    if path_etat_prec and os.path.exists(path_etat_prec):
        print(f"Traitement de l'état précédent : {path_etat_prec}")
        try:
            # Lecture brute sans header
            df_etat = pd.read_excel(path_etat_prec, header=None)
            
            start_idx = 3 # Ligne 4
            # Modification : On traite de la ligne 4 jusqu'à l'avant-avant-dernière (exclusion des 3 lignes de fin)
            if len(df_etat) > start_idx:
                for idx, row in df_etat.iloc[start_idx:-3].iterrows():
                    # Arrêt si ligne de totaux/fin (Sécurité anti-bug)
                    val_lib = str(row[1]) if pd.notna(row[1]) else ""
                    if any(x in val_lib.lower() for x in ["total", "totaux", "solde"]):
                        continue
                    
                    # On s'assure que row a assez de colonnes (au moins 6 pour aller jusqu'à F)
                    if len(row) < 6: continue
                    
                    # Lecture des montants
                    val_c = row[2] if isinstance(row[2], (int, float)) else 0 # Debit Compta
                    val_d = row[3] if isinstance(row[3], (int, float)) else 0 # Credit Compta
                    val_e = row[4] if isinstance(row[4], (int, float)) else 0 # Debit Banque
                    val_f = row[5] if isinstance(row[5], (int, float)) else 0 # Credit Banque
                    
                    keep_c, keep_d, keep_e, keep_f = 0, 0, 0, 0
                    
                    # 1. Verification Compta (Col C/D vs Journal)
                    # Col C (Debit) -> Chercher dans Debit Journal
                    if val_c > 0:
                        match = df_compta[
                            (df_compta['debit'] == val_c) & 
                            (~df_compta.index.isin(indices_compta_ok))
                        ]
                        if not match.empty:
                            indices_compta_ok.append(match.index[0])
                            # Trouvé -> Eliminé (ne rien faire)
                        else:
                            keep_c = val_c # Pas trouvé -> A reporter

                    # Col D (Credit) -> Chercher dans Credit Journal
                    if val_d > 0:
                        match = df_compta[
                            (df_compta['credit'] == val_d) & 
                            (~df_compta.index.isin(indices_compta_ok))
                        ]
                        if not match.empty:
                            indices_compta_ok.append(match.index[0])
                        else:
                            keep_d = val_d

                    # 2. Verification Banque (Col E/F vs Relevé)
                    # Col E (Debit) -> Chercher Débit Relevé
                    if val_e > 0:
                        match = df_banque[
                            (df_banque['debit'] == val_e) & 
                            (~df_banque.index.isin(indices_banque_ok))
                        ]
                        if not match.empty:
                            indices_banque_ok.append(match.index[0])
                        else:
                            keep_e = val_e
                            
                    # Col F (Credit) -> Chercher Crédit Relevé
                    if val_f > 0:
                        match = df_banque[
                            (df_banque['credit'] == val_f) & 
                            (~df_banque.index.isin(indices_banque_ok))
                        ]
                        if not match.empty:
                            indices_banque_ok.append(match.index[0])
                        else:
                            keep_f = val_f
                    
                    # Si au moins un montant est à reporter
                    if any([keep_c, keep_d, keep_e, keep_f]):
                        suspens_etat_prec.append({
                            'raw_date': get_date_obj(row[0]),
                            'date_str': format_date_val(row[0]),
                            'libelle': row[1],
                            'col_C': keep_c,
                            'col_D': keep_d,
                            'col_E': keep_e,
                            'col_F': keep_f
                        })

        except Exception as e:
            print(f"Erreur lors de la lecture de l'état précédent : {e}")

    # --- LOGIQUE DE POINTAGE ---
    # On parcourt le relevé bancaire
    for idx_b, row_b in df_banque.iterrows():
        
        # Cas 1 : Opération au débit de la banque -> On cherche au crédit de la compta
        if row_b['debit'] > 0:
            match = df_compta[
                (df_compta['credit'] == row_b['debit']) & 
                (~df_compta.index.isin(indices_compta_ok))
            ]
            if not match.empty:
                indices_banque_ok.append(idx_b)
                indices_compta_ok.append(match.index[0])
                
        # Cas 2 : Opération au crédit de la banque -> On cherche au débit de la compta
        elif row_b['credit'] > 0:
            match = df_compta[
                (df_compta['debit'] == row_b['credit']) & 
                (~df_compta.index.isin(indices_compta_ok))
            ]
            if not match.empty:
                indices_banque_ok.append(idx_b)
                indices_compta_ok.append(match.index[0])

    # --- EXTRACTION DES SUSPENS (Non trouvés) ---
    suspens_banque = df_banque.drop(indices_banque_ok)
    suspens_compta = df_compta.drop(indices_compta_ok)

    output_file = output_path

    # 1. Exportation des suspens vers le fichier Excel (crée le fichier)
    # 1. Exportation des suspens vers le fichier Excel (crée le fichier)
    print(f"Création du fichier {output_file}...")
    
    # Suppression de la colonne Solde pour l'export
    # On cherche les colonnes qui contiennent "solde" (indépendamment de la casse)
    cols_to_drop_banque = [c for c in suspens_banque.columns if 'solde' in str(c).lower()]
    suspens_banque_export = suspens_banque.drop(columns=cols_to_drop_banque)
    
    cols_to_drop_compta = [c for c in suspens_compta.columns if 'solde' in str(c).lower()]
    suspens_compta_export = suspens_compta.drop(columns=cols_to_drop_compta)

    # Nettoyage de la date pour le journal (enlever l'heure)
    col_date_compta = next((c for c in suspens_compta_export.columns if 'date' in str(c).lower()), None)
    if col_date_compta:
        try:
            suspens_compta_export[col_date_compta] = pd.to_datetime(suspens_compta_export[col_date_compta], errors='coerce').dt.date
        except Exception as e:
            print(f"Attention: Impossible de formater la date pour le journal : {e}")

    with pd.ExcelWriter(output_file) as writer:
        suspens_banque_export.to_excel(writer, sheet_name='RELEVE_NON_POINTEE', index=False)
        suspens_compta_export.to_excel(writer, sheet_name='JOURNAL_NON_POINTEE', index=False)

    print(f"Analyse terminée.")
    print(f"- Opérations banque en suspens : {len(suspens_banque)}")
    print(f"- Opérations compta en suspens : {len(suspens_compta)}")

    # 2. Ajout et formatage de la feuille RAPPROCHEMENT
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        
        # Récupération des soldes (insensible à la casse)
        def get_last_solde(df, file_label):
            col_solde = next((c for c in df.columns if str(c).lower().strip() == 'solde'), None)
            if col_solde:
                return df[col_solde].iloc[-1]
            return 0

        # On relit les fichiers sources pour les soldes
        df_banque_raw = pd.read_excel(path_banque)
        solde_banque = get_last_solde(df_banque_raw, "banque")

        df_compta_raw = pd.read_excel(path_compta, header=1)
        solde_compta = get_last_solde(df_compta_raw, "compta")

        # Chargement du classeur de sortie pour ajout de la feuille
        wb = openpyxl.load_workbook(output_file)
        
        sheet_name_rapp = "RAPPROCHEMENT"
        if sheet_name_rapp in wb.sheetnames:
            del wb[sheet_name_rapp]
        ws = wb.create_sheet(sheet_name_rapp)

        # --- Création du tableau RAPPROCHEMENT ---
        # En-têtes
        ws['A1'] = "Date"
        ws['B1'] = "Libellés"
        ws['C1'] = "Compte courant"
        ws['E1'] = "Relevé bancaire"
        ws['C2'] = "Débit"
        ws['D2'] = "Crédit"
        ws['E2'] = "Débit"
        ws['F2'] = "Crédit"

        # Contenu ligne 3
        ws['B3'] = "Solde à rectifier"
        ws['C3'] = solde_compta  # Solde compta
        ws['F3'] = solde_banque  # Solde banque (crédit ? vérifer emplacement usuel, user a demandé F3 qui est Crédit Banque)

        # Fusion
        ws.merge_cells('A1:A2')
        ws.merge_cells('B1:B2')
        ws.merge_cells('C1:D1')
        ws.merge_cells('E1:F1')

        # Styles
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Style Header (Lignes 1-2)
        for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=6):
            for cell in row:
                cell.font = bold_font
                cell.alignment = center_align
                cell.border = thin_border

        # Style Ligne 3
        for col in range(1, 7):
            cell = ws.cell(row=3, column=col)
            cell.border = thin_border
            if col == 2:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = center_align

        # Largeur colonnes
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15

        # Helpers dates (définis au début de la fonction)


        # Identification des colonnes
        col_date_c = next((c for c in suspens_compta.columns if 'date' in str(c).lower()), None)
        col_lib_c = next((c for c in suspens_compta.columns if 'libell' in str(c).lower()), None)
        
        col_date_b = next((c for c in suspens_banque.columns if 'date' in str(c).lower()), None)
        col_lib_b = next((c for c in suspens_banque.columns if 'libell' in str(c).lower()), None)

        # 1. Collecte des opérations dans une liste commune
        all_ops = []
        
        # --- Ajout des reports de l'ETAT PRECEDENT ---
        for item in suspens_etat_prec:
            # On n'ajoute la ligne que si elle contient quelque chose (sécurité)
            if any([item['col_C'], item['col_D'], item['col_E'], item['col_F']]):
                all_ops.append({
                    'raw_date': item['raw_date'],
                    'date_str': item['date_str'],
                    'libelle': item['libelle'],
                    'col_C': item['col_C'], 
                    'col_D': item['col_D'], 
                    'col_E': item['col_E'], 
                    'col_F': item['col_F'] 
                })

        # --- Opérations du JOURNAL (Compta) ---
        # Debit Compta -> Crédit Relevé (F)
        # Crédit Compta -> Débit Relevé (E)
        for _, row in suspens_compta.iterrows():
            d_val = row.get('debit', 0)
            c_val = row.get('credit', 0)
            
            raw_date = row[col_date_c] if col_date_c else None
            libelle = row[col_lib_c] if col_lib_c else ""
            
            if d_val > 0:
                all_ops.append({
                    'raw_date': get_date_obj(raw_date),
                    'date_str': format_date_val(raw_date),
                    'libelle': libelle,
                    'col_C': 0, 'col_D': 0, 'col_E': 0, 'col_F': d_val 
                })
            elif c_val > 0:
                all_ops.append({
                    'raw_date': get_date_obj(raw_date),
                    'date_str': format_date_val(raw_date),
                    'libelle': libelle,
                    'col_C': 0, 'col_D': 0, 'col_E': c_val, 'col_F': 0
                })

        # --- Opérations du RELEVE (Banque) ---
        # Debit Banque -> Crédit Compte Courant (D)
        # Crédit Banque -> Débit Compte Courant (C)
        for _, row in suspens_banque.iterrows():
            d_val = row.get('debit', 0)
            c_val = row.get('credit', 0)
            
            raw_date_b = row[col_date_b] if col_date_b else None
            libelle_b = row[col_lib_b] if col_lib_b else ""
            
            if d_val > 0:
                all_ops.append({
                    'raw_date': get_date_obj(raw_date_b),
                    'date_str': format_date_val(raw_date_b),
                    'libelle': libelle_b,
                    'col_C': 0, 'col_D': d_val, 'col_E': 0, 'col_F': 0
                })
            elif c_val > 0:
                all_ops.append({
                    'raw_date': get_date_obj(raw_date_b),
                    'date_str': format_date_val(raw_date_b),
                    'libelle': libelle_b,
                    'col_C': c_val, 'col_D': 0, 'col_E': 0, 'col_F': 0
                })

        # 2. Tri des opérations par date
        all_ops.sort(key=lambda x: x['raw_date'])

        # 3. Écriture dans le tableau
        current_row = 4
        
        # Si aucune opération en suspens (tout est pointé), on ajoute une ligne vierge pour l'esthétique
        if not all_ops:
             for col in range(1, 7):
                 cell = ws.cell(row=current_row, column=col)
                 cell.border = thin_border
             current_row += 1

        for op in all_ops:
            ws[f'A{current_row}'] = op['date_str']
            ws[f'B{current_row}'] = op['libelle']
            if op['col_C'] > 0: ws[f'C{current_row}'] = op['col_C']
            if op['col_D'] > 0: ws[f'D{current_row}'] = op['col_D']
            if op['col_E'] > 0: ws[f'E{current_row}'] = op['col_E']
            if op['col_F'] > 0: ws[f'F{current_row}'] = op['col_F']
            
            # Styles
            for col in range(1, 7):
                cell = ws.cell(row=current_row, column=col)
                cell.border = thin_border
                if col == 2:
                        cell.alignment = Alignment(horizontal='left')
                else:
                        cell.alignment = center_align
            current_row += 1

        # --- LIGNE TOTAL ---
        ws[f'B{current_row}'] = "Totaux"
        ws[f'B{current_row}'].font = bold_font
        ws[f'B{current_row}'].alignment = Alignment(horizontal='center')
        ws[f'B{current_row}'].border = thin_border
        
        # Application des formules de somme et styles sur C, D, E, F
        for col in ['C', 'D', 'E', 'F']:
            cell = ws[f'{col}{current_row}']
            cell.value = f"=SUM({col}3:{col}{current_row-1})"
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = center_align
            
        # On applique aussi la bordure sur la colonne A pour finir proprement le tableau
        ws[f'A{current_row}'].border = thin_border

        row_total = current_row
        current_row += 1

        # --- LIGNE SOLDE RECTIFIE ---
        label_rectif = "Solde rectifié"
        if date_rapprochement:
             # On formate la date si c'est un objet date/datetime, sinon on l'utilise telle quelle
            try:
                date_str = date_rapprochement.strftime('%d/%m/%Y')
            except AttributeError:
                date_str = str(date_rapprochement)
            label_rectif = f"Solde rectifié au {date_str}"
            
        ws[f'B{current_row}'] = label_rectif
        ws[f'B{current_row}'].font = bold_font
        ws[f'B{current_row}'].alignment = Alignment(horizontal='center') # Alignement centré comme demandé
        ws[f'B{current_row}'].border = thin_border

        # Logique COMPTE COURANT (Col C = Débit, Col D = Crédit)
        # Règle user : "total débit - total crédit. si le montant est positif, alors l'inscrire dans la colonne credit"
        # On ajoute aussi le cas inverse pour l'équilibre (si solde créditeur, on le met au débit)
        
        # Cellule C (Débit) : SI(Crédit > Débit; Crédit - Débit; "")
        ws[f'C{current_row}'] = f'=IF(D{row_total}-C{row_total}>0, D{row_total}-C{row_total}, "")'
        
        # Cellule D (Crédit) : SI(Débit > Crédit; Débit - Crédit; "")
        ws[f'D{current_row}'] = f'=IF(C{row_total}-D{row_total}>0, C{row_total}-D{row_total}, "")'

        # Logique RELEVE BANCAIRE (Col E = Débit, Col F = Crédit)
        # Règle "sens inverse" (= même logique d'équilibrage) :
        # - Si Crédit > Débit, on met la différence au Débit (E)
        # - Si Débit > Crédit, on met la différence au Crédit (F)
        
        # Cellule E (Débit) : SI(Crédit > Débit; Crédit - Débit; "")
        ws[f'E{current_row}'] = f'=IF(F{row_total}-E{row_total}>0, F{row_total}-E{row_total}, "")'
        
        # Cellule F (Crédit) : SI(Débit > Crédit; Débit - Crédit; "")
        ws[f'F{current_row}'] = f'=IF(E{row_total}-F{row_total}>0, E{row_total}-F{row_total}, "")'

        # Styles de la ligne Solde Rectifié
        for col in ['A', 'C', 'D', 'E', 'F']:
            cell = ws[f'{col}{current_row}']
            cell.border = thin_border
            cell.alignment = center_align
            cell.font = bold_font

        row_rectifie = current_row
        current_row += 1

        # --- LIGNE TOTAUX GENERAUX ---
        ws[f'B{current_row}'] = "TOTAUX GENERAUX"
        ws[f'B{current_row}'].font = bold_font
        ws[f'B{current_row}'].alignment = Alignment(horizontal='center')
        ws[f'B{current_row}'].border = thin_border
        
        for col in ['C', 'D', 'E', 'F']:
            cell = ws[f'{col}{current_row}']
            # Somme de la ligne Totaux et Solde rectifié
            # On utilise SUM pour éviter les erreurs si une cellule est vide ("")
            cell.value = f"=SUM({col}{row_total}, {col}{row_rectifie})"
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = center_align
            
        ws[f'A{current_row}'].border = thin_border

        # --- FORMATAGE COMPTABLE ---
        # On applique le format '#,##0' sur toutes les cellules de montants (C, D, E, F) de la ligne 3 à la fin
        # Cela permet d'avoir les séparateurs de milliers et 0 décimales (entiers)
        accounting_format = '#,##0'
        
        for row in range(3, current_row + 1):
            for col in ['C', 'D', 'E', 'F']:
                cell = ws[f'{col}{row}']
                cell.number_format = accounting_format

        wb.save(output_file)
        print(f"Feuille {sheet_name_rapp} ajoutée et mise à jour dans {output_file}.")

    except Exception as e:
        print(f"Erreur lors de la création de la feuille Rapprochement : {e}")

# Pour lancer l'outil :
if __name__ == "__main__":
    executer_rapprochement('documents/transactions_globales.xlsx', 'documents/journal_banque.xlsx')